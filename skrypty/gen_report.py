# -*- coding: utf-8 -*-
"""Raport XLS: które firmy NIE mają strony / Facebooka / LinkedIn (+ NIP/REGON/KRS jeśli z GUS).
Output: gig.org.pl/raport_czlonkowie_GIG.xlsx"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
recs = json.load(open(os.path.join(HERE, "_czlonkowie.json"), encoding="utf-8"))
enr = {e["name"]: e for e in (json.load(open(os.path.join(HERE, "_czlonkowie_enrich.json"), encoding="utf-8")) if os.path.exists(os.path.join(HERE, "_czlonkowie_enrich.json")) else [])}
spath = os.path.join(HERE, "_czlonkowie_social.json")
soc = json.load(open(spath, encoding="utf-8")) if os.path.exists(spath) else {}
gpath = os.path.join(HERE, "_czlonkowie_ids_gus.json")
gus = json.load(open(gpath, encoding="utf-8")) if os.path.exists(gpath) else {}
FREE = {"gmail.com","wp.pl","o2.pl","interia.pl","interia.eu","onet.pl","op.pl","poczta.fm","vp.pl",
        "poczta.onet.pl","ko.onet.pl","gazeta.pl","outlook.com","hotmail.com","yahoo.com","tlen.pl","go2.pl"}
def clean(v): return "" if (not v or str(v).strip().lower()=="brak") else str(v).strip()
def website(r):
    em=(r.get("email") or "").lower()
    if "@" in em:
        d=em.split("@")[-1]
        if d and d not in FREE and "." in d: return "https://"+d
    w = clean(soc.get(r["name"], {}).get("website")) or clean(enr.get(r["name"], {}).get("website"))
    return w
def fb_of(r): return clean(soc.get(r["name"],{}).get("facebook")) or clean(enr.get(r["name"],{}).get("facebook"))
def li_of(r): return clean(soc.get(r["name"],{}).get("linkedin")) or clean(enr.get(r["name"],{}).get("linkedin"))
def city(a):
    m=re.search(r"\d{2}-\d{3}\s*(.+)$", a or ""); return m.group(1).strip() if m else ""

wb=Workbook(); ws=wb.active; ws.title="Członkowie GIG"
hdr=["Lp.","Województwo","Firma","Miejscowość","E-mail","Strona WWW","Ma stronę?",
     "Facebook","Ma FB?","LinkedIn","Ma LinkedIn?","NIP","REGON","KRS"]
red=PatternFill("solid",fgColor="F8C9C9"); grn=PatternFill("solid",fgColor="CDEBCD")
head=PatternFill("solid",fgColor="CC0A2B"); thin=Side(style="thin",color="DDDDDD")
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ws.append(hdr)
for c in ws[1]:
    c.font=Font(bold=True,color="FFFFFF"); c.fill=head; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bd
n_www=n_fb=n_li=0
for i,r in enumerate(recs,1):
    g=gus.get(r["name"],{})
    www=website(r); fb=fb_of(r); li=li_of(r)
    has_www="TAK" if www else "NIE"; has_fb="TAK" if fb else "NIE"; has_li="TAK" if li else "NIE"
    if not www: n_www+=1
    if not fb: n_fb+=1
    if not li: n_li+=1
    nip=g.get("nip") if g.get("gus") else ""; regon=g.get("regon") if g.get("gus") else ""; krs=g.get("krs") if g.get("gus") else ""
    row=[i, r.get("region",""), r["name"], city(r.get("address","")), r.get("email",""),
         www or "brak", has_www, fb or "brak", has_fb, li or "brak", has_li, nip or "", regon or "", krs or ""]
    ws.append(row)
    rr=ws[ws.max_row]
    for c in rr: c.border=bd; c.alignment=Alignment(vertical="center",wrap_text=(c.column in (3,6,8,10)))
    rr[6].fill = grn if www else red   # Ma stronę?
    rr[8].fill = grn if fb else red    # Ma FB?
    rr[10].fill = grn if li else red   # Ma LinkedIn?
    for idx in (6,8,10): rr[idx].alignment=Alignment(horizontal="center")
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:N{ws.max_row}"
widths=[5,16,42,18,26,30,10,30,9,30,12,13,12,12]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w

# Arkusz podsumowania
s=wb.create_sheet("Podsumowanie")
s.append(["Podsumowanie braków — Członkowie GIG"]); s["A1"].font=Font(bold=True,size=14,color="CC0A2B")
s.append([]); s.append(["Liczba firm ogółem", len(recs)])
s.append(["Bez strony WWW", n_www]); s.append(["Bez profilu Facebook", n_fb]); s.append(["Bez profilu LinkedIn", n_li])
s.append([]); s.append(["Ze stroną WWW", len(recs)-n_www]); s.append(["Z Facebookiem", len(recs)-n_fb]); s.append(["Z LinkedIn", len(recs)-n_li])
if not gus: s.append([]); s.append(["Uwaga:", "NIP/REGON/KRS uzupełnią się po pobraniu z GUS (klient gus_fetch.py)."])
for row in s.iter_rows(min_row=3):
    if row and row[0].value is not None: row[0].font=Font(bold=True)
s.column_dimensions["A"].width=28; s.column_dimensions["B"].width=60

outp=os.path.join(ROOT, "raport_czlonkowie_GIG.xlsx")
wb.save(outp)
print("Zapisano:", outp)
print(f"Bez strony: {n_www} | bez FB: {n_fb} | bez LinkedIn: {n_li} | firm: {len(recs)}")
